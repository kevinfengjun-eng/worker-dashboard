#!/usr/bin/env python3
"""
临时工管理看板 - 数据预嵌入生成脚本
从Excel台账读取数据，嵌入到worker_dashboard.html中，生成自包含的静态HTML。
用法: python3 generate_embedded.py [Excel文件] [输出HTML]
"""
import sys, os, json, re
from collections import Counter, defaultdict
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)

# ---- 默认路径 ----
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = sys.argv[1] if len(sys.argv) > 1 else os.path.join(WORK_DIR, '临时工管理台账02.xlsx')
OUTPUT_FILE = sys.argv[2] if len(sys.argv) > 2 else os.path.join(WORK_DIR, 'worker_dashboard_embedded.html')
TEMPLATE_FILE = os.path.join(WORK_DIR, 'worker_dashboard.html')

def to_date(v):
    if isinstance(v, datetime): return v
    if isinstance(v, (int, float)) and v > 40000:
        return datetime(1899, 12, 30) + timedelta(days=int(v))
    return None

def main():
    print(f"读取Excel: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # 供应商名称
    ws_sp = wb['供应商管理']
    supplier_names = []
    for row in ws_sp.iter_rows(min_row=2, max_row=ws_sp.max_row, values_only=True):
        if row[2]:
            supplier_names.append(str(row[2]).strip())

    # 在职清单
    ws_act = wb['在职清单']
    active_data = []
    for row in ws_act.iter_rows(min_row=2, max_row=ws_act.max_row, values_only=True):
        if row[0]:
            active_data.append({
                'job': str(row[3] or '').strip(),
                'skill': str(row[4] or '').strip(),
                'supplier': str(row[5] or '').strip(),
                'entry_date': row[6],
                'region': str(row[8] or '').strip(),
            })

    # 面试清单
    ws_intv = wb['面试清单']
    intv_data = []
    for row in ws_intv.iter_rows(min_row=2, max_row=ws_intv.max_row, values_only=True):
        if row[0]:
            intv_data.append({
                'date': str(row[1] or '').strip(),
                'name': str(row[2] or '').strip(),
                'job': str(row[4] or '').strip(),
                'skill': str(row[5] or '').strip(),
                'supplier': str(row[6] or '').strip(),
                'channel': str(row[7] or '').strip(),
                'result': str(row[8] or '').strip(),
                'reported': str(row[9] or '').strip(),
                'region': str(row[13] or '').strip(),
            })

    # 离职清单
    ws_lv = wb['离职清单']
    leave_data = []
    for row in ws_lv.iter_rows(min_row=2, max_row=ws_lv.max_row, values_only=True):
        if row[0]:
            leave_data.append({
                'name': str(row[1] or '').strip(),
                'job': str(row[3] or '').strip(),
                'skill': str(row[4] or '').strip(),
                'supplier': str(row[5] or '').strip(),
                'join_date': row[6],
                'leave_date': row[7],
                'days': str(row[8] or '').strip(),
                'reason': str(row[9] or '').strip(),
                'reason_detail': str(row[10] or '').strip(),
                'settled': str(row[11] or '').strip(),
                'region': str(row[13] or '').strip(),
            })

    # KPI
    interview_total = len(intv_data)
    interview_pass = sum(1 for r in intv_data if r['result'] == '通过')
    report_count = sum(1 for r in intv_data if r['reported'] == '是')

    # 统计
    job_names = ['电工', '钳工', '调试', '仓库', '其他']
    skill_names = ['大工', '中工', '小工']
    jm_active = 0; cd_active = 0
    jm_jobs = {}; cd_jobs = {}
    jm_skills = {}; cd_skills = {}; skills = {}
    jm_job_skills = {}; cd_job_skills = {}
    for j in job_names:
        jm_jobs[j] = 0; cd_jobs[j] = 0
        jm_job_skills[j] = {}; cd_job_skills[j] = {}
        for s in skill_names:
            jm_job_skills[j][s] = 0; cd_job_skills[j][s] = 0
    for s in skill_names:
        skills[s] = 0; jm_skills[s] = 0; cd_skills[s] = 0

    supplier_count = {n: 0 for n in supplier_names}
    total_active = 0

    for r in active_data:
        if not r['region']: continue
        total_active += 1
        is_jm = '江门' in r['region']
        if is_jm: jm_active += 1
        else: cd_active += 1
        j, s, sp = r['job'], r['skill'], r['supplier']
        if j in jm_jobs: (jm_jobs if is_jm else cd_jobs)[j] += 1
        if s in skills:
            skills[s] += 1
            (jm_skills if is_jm else cd_skills)[s] += 1
        if j in jm_job_skills and s in skill_names:
            (jm_job_skills if is_jm else cd_job_skills)[j][s] += 1
        if sp in supplier_count: supplier_count[sp] += 1
        elif sp:
            supplier_count[sp] = 1
            supplier_names.append(sp)

    sorted_suppliers = sorted([n for n in supplier_names if supplier_count.get(n, 0) > 0], key=lambda n: supplier_count[n], reverse=True)
    top_suppliers = sorted_suppliers[:15]

    # 入职/离职 按日期
    join_by_date_region = defaultdict(lambda: {'jm': 0, 'cd': 0})
    join_by_date_job = defaultdict(lambda: {'jm': {}, 'cd': {}})
    month_start = datetime(datetime.now().year, datetime.now().month, 1)
    jm_monthly_join = 0; cd_monthly_join = 0

    for r in active_data:
        d = to_date(r['entry_date'])
        if not d or not r['region']: continue
        ds = d.strftime('%Y-%m-%d')
        rk = 'jm' if '江门' in r['region'] else 'cd'
        join_by_date_region[ds][rk] += 1
        j = r['job']
        join_by_date_job[ds][rk][j] = join_by_date_job[ds][rk].get(j, 0) + 1
        if d >= month_start:
            if rk == 'jm': jm_monthly_join += 1
            else: cd_monthly_join += 1

    leave_by_date_region = defaultdict(lambda: {'jm': 0, 'cd': 0})
    leave_by_date_job = defaultdict(lambda: {'jm': {}, 'cd': {}})
    jm_monthly_leave = 0; cd_monthly_leave = 0

    for r in leave_data:
        d = to_date(r['leave_date'])
        if not d or not r['region']: continue
        ds = d.strftime('%Y-%m-%d')
        rk = 'jm' if '江门' in r['region'] else 'cd'
        leave_by_date_region[ds][rk] += 1
        j = r['job']
        leave_by_date_job[ds][rk][j] = leave_by_date_job[ds][rk].get(j, 0) + 1
        if d >= month_start:
            if rk == 'jm': jm_monthly_leave += 1
            else: cd_monthly_leave += 1

    # 日报
    all_dates = sorted(set(list(join_by_date_region.keys()) + list(leave_by_date_region.keys())))
    if not all_dates:
        print("错误: 未找到有效日期数据"); sys.exit(1)
    start_d = datetime.strptime(all_dates[0], '%Y-%m-%d')
    end_d = datetime.now()
    day = timedelta(days=1)

    daily = []
    cum_jm = 0; cum_cd = 0
    d = start_d
    while d <= end_d:
        ds = d.strftime('%Y-%m-%d')
        jj = join_by_date_region.get(ds, {}).get('jm', 0)
        jl = leave_by_date_region.get(ds, {}).get('jm', 0)
        cj = join_by_date_region.get(ds, {}).get('cd', 0)
        cl = leave_by_date_region.get(ds, {}).get('cd', 0)
        cum_jm += jj - jl
        cum_cd += cj - cl
        daily.append({
            'date': f"{d.month:02d}-{d.day:02d}",
            'dateObj_ts': int(d.timestamp() * 1000),
            'jm_j': jj, 'jm_l': jl, 'jm_n': jj - jl,
            'cd_j': cj, 'cd_l': cl, 'cd_n': cj - cl,
            'cumJm': cum_jm, 'cumCd': cum_cd,
            'jobs_jm': [join_by_date_job.get(ds, {}).get('jm', {}).get(j, 0) for j in job_names],
            'jobs_cd': [join_by_date_job.get(ds, {}).get('cd', {}).get(j, 0) for j in job_names],
        })
        d += day

    last7 = daily[-7:]
    today = datetime.now()

    DATA = {
        'reportDate': today.strftime('%Y-%m-%d'),
        'periodStart': start_d.strftime('%Y-%m-%d'),
        '_isRealData': True,
        'targets': {'江门': 2500, '成都': 900},
        'skillDemand': {'大工': 450, '中工': 1500, '小工': 2000},
        'snapshot': {
            'totalActive': total_active, 'jmActive': jm_active, 'cdActive': cd_active,
            'jmJobs': jm_jobs, 'cdJobs': cd_jobs,
            'jmSkills': jm_skills, 'cdSkills': cd_skills,
            'jmJobSkills': jm_job_skills, 'cdJobSkills': cd_job_skills,
            'skills': skills,
            'supplierCount': {n: supplier_count.get(n, 0) for n in top_suppliers},
            'topSuppliers': top_suppliers, 'supplierNames': top_suppliers,
            'interviewTotal': interview_total, 'interviewPass': interview_pass, 'reportCount': report_count,
        },
        'weekly': {
            'jmJoin': sum(d['jm_j'] for d in last7), 'jmLeave': sum(d['jm_l'] for d in last7),
            'cdJoin': sum(d['cd_j'] for d in last7), 'cdLeave': sum(d['cd_l'] for d in last7),
        },
        'monthly': {
            'jmJoin': jm_monthly_join, 'jmLeave': jm_monthly_leave,
            'cdJoin': cd_monthly_join, 'cdLeave': cd_monthly_leave,
        },
        'prevWeekActive': total_active - sum(d['jm_n'] + d['cd_n'] for d in last7),
        'prevWeekJmActive': jm_active - sum(d['jm_n'] for d in last7),
        'prevWeekCdActive': cd_active - sum(d['cd_n'] for d in last7),
        'daily': daily,
        # 面试数据
        'interview': {
            'channels': dict(sorted(
                {ch: sum(1 for r in intv_data if r['channel'] == ch) for ch in set(r['channel'] for r in intv_data if r['channel'])}.items(),
                key=lambda x: -x[1]
            )),
            'regions': dict(sorted(
                {rg: sum(1 for r in intv_data if r['region'] == rg) for rg in set(r['region'] for r in intv_data if r['region'])}.items(),
                key=lambda x: -x[1]
            )),
            'suppliers': dict(sorted(
                {sp: sum(1 for r in intv_data if r['supplier'] == sp) for sp in set(r['supplier'] for r in intv_data if r['supplier'])}.items(),
                key=lambda x: -x[1]
            )),
        },
        # 离职数据
        'resign': {
            'total': len([r for r in leave_data if r['region']]),
            'settled': sum(1 for r in leave_data if r['settled'] == '是'),
            'reasons': dict(sorted(
                {rs: sum(1 for r in leave_data if r['reason'] == rs) for rs in set(r['reason'] for r in leave_data if r['reason'])}.items(),
                key=lambda x: -x[1]
            )),
            'jobs': dict(sorted(
                {j: sum(1 for r in leave_data if r['job'] == j) for j in set(r['job'] for r in leave_data if r['job'])}.items(),
                key=lambda x: -x[1]
            )),
            'suppliers': dict(sorted(
                {sp: sum(1 for r in leave_data if r['supplier'] == sp) for sp in set(r['supplier'] for r in leave_data if r['supplier'])}.items(),
                key=lambda x: -x[1]
            )),
            'details': [
                {
                    'name': r['name'],
                    'job': r['job'],
                    'supplier': r['supplier'],
                    'joinDate': str(to_date(r['join_date']).strftime('%Y-%m-%d')) if to_date(r['join_date']) else '',
                    'leaveDate': str(to_date(r['leave_date']).strftime('%Y-%m-%d')) if to_date(r['leave_date']) else '',
                    'days': r['days'],
                    'reason': r['reason'],
                    'region': r['region'],
                }
                for r in leave_data if r['region']
            ],
        },
    }

    # 读取模板HTML
    print(f"读取模板: {TEMPLATE_FILE}")
    with open(TEMPLATE_FILE, 'r', encoding='utf-8') as f:
        html = f.read()

    data_json = json.dumps(DATA, ensure_ascii=False, separators=(',', ':'))

    embedded_js = f'''
// ===== EMBEDDED DATA (预嵌入真实数据) =====
const EMBEDDED_DATA = {data_json};

function initWithEmbeddedData(){{
  destroyAllCharts();
  if(EMBEDDED_DATA.daily){{EMBEDDED_DATA.daily.forEach(d=>{{if(d.dateObj_ts)d.dateObj=new Date(d.dateObj_ts);delete d.dateObj_ts}})}};
  Object.assign(DATA, JSON.parse(JSON.stringify(EMBEDDED_DATA)));
  // 重建dateObj（JSON.stringify会把Date变成ISO字符串）
  DATA.daily.forEach(d=>{{
    if(d.dateObj && typeof d.dateObj==='string'){{
      d.dateObj=new Date(d.dateObj);
    }} else if(!d.dateObj && d.date){{
      const parts=d.date.split('-');
      if(parts.length===2) d.dateObj=new Date(2026,parseInt(parts[0])-1,parseInt(parts[1]));
      else if(parts.length===3) d.dateObj=new Date(d.date);
    }}
    if(!(d.dateObj instanceof Date)||isNaN(d.dateObj.getTime())) d.dateObj=new Date();
  }});
  DATA._isRealData=true;
  if(DATA.snapshot.topSuppliers){{COLORS.supplierNames=DATA.snapshot.topSuppliers;const p=['#3B82F6','#22C55E','#F59E0B','#8B5CF6','#EC4899','#14B8A6','#F97316','#6366F1','#EF4444','#84CC16','#06B6D4','#D946EF','#0EA5E9','#A855F7','#F43F5E'];COLORS.suppliers=DATA.snapshot.topSuppliers.map((_,i)=>p[i%p.length])}};
  ensureDefaults();
  $('#headerSub').innerHTML='真实数据（预嵌入） · 报告周期：'+DATA.periodStart+' ~ '+DATA.reportDate;
  $('#reportDate').value=DATA.reportDate;
  setupDatePickers();initTrendControls();
  renderOverviewTab();_renderedTabs.add('overview');
}}

function doLogin(){{CURRENT_ROLE='viewer';$('#loginOverlay').style.display='none';$('#mainApp').style.display='block';applyRoleRestriction();initWithEmbeddedData()}}

window.addEventListener('DOMContentLoaded',function(){{setTimeout(function(){{$('#loginOverlay').style.display='none';$('#mainApp').style.display='block';applyRoleRestriction();initWithEmbeddedData()}},100)}});
'''

    # 替换模板中的DATA声明
    html = html.replace('let DATA = {};', f'let DATA = {{}};\n\n{embedded_js}')

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"\n生成完成: {OUTPUT_FILE}")
    print(f"在职人数: {total_active} (江门{jm_active} + 成都{cd_active})")
    print(f"面试: {interview_total}人, 通过: {interview_pass}人, 报到: {report_count}人")
    print(f"日报天数: {len(daily)}")

if __name__ == '__main__':
    main()
